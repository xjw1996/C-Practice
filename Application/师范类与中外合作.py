
import os
import csv
from importlib.resources import path
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

path = r'D:\\xue_workspace\\DataOfUniversityScore\\ExcelPythonTest\\专业\\师范类与中外合作理科一批'
path_="D:\\xue_workspace\\DataOfUniversityScore\\ExcelPythonTest\\专业\\师范类与中外合作理科一批\\"


filenames = os.listdir(path)
for filename_ in filenames:
    str_arr = filename_.split('.')
    #去掉csv后缀的名字
    name_=str_arr[0]
    xlsx_filename = "D:\\xue_workspace\\DataOfUniversityScore\\ExcelPythonTest\\专业\\师范类与中外合作理科一批\\"+name_+".xlsx"
    csv_filename=path_+filename_
    font = Font(name='宋体',size=18)
    wb = openpyxl.Workbook()
    wb.save("D:\\xue_workspace\\DataOfUniversityScore\\ExcelPythonTest\\专业\\师范类与中外合作理科一批\\"+name_+".xlsx")
    with open(csv_filename,newline="") as csvf:
        #CSVファイルを読み込む
        data=csv.reader(csvf)
        #Excelファイルを開く
        wb=openpyxl.load_workbook(xlsx_filename)
        #sheetを読み込む
        ws=wb["Sheet"]
        ws['A1'] = name_
        ws['A1'].font = font
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells(range_string="A1:k1")
        ws['A2'] ='院校代码'
        ws['B2'] ='院校'
        ws['C2'] ='备注专业代码'
        ws['D2'] ='专业'
        ws['E2'] ='专业备注'
        ws['F2'] ='最低分'
        ws['G2'] ='最低位次'
        ws['H2'] ='录取数'
        ws['I2'] ='性质'
        ws['J2'] ='水平'
        ws['k2'] ='类型'
        r=3
        for line in data: 
            c=1
            for v in line:
                ws.cell(row=r,column=c).value=v
                c += 1
            r += 1
        for num in range(2, r):
            A_colom="A"+str(num)
            F_colom="F"+str(num)
            G_colom="G"+str(num)
            H_colom="H"+str(num)
            I_colom="I"+str(num)
            J_colom="J"+str(num)
            K_colom="K"+str(num)
            ws[A_colom].alignment = Alignment(horizontal="center")
            ws[F_colom].alignment = Alignment(horizontal="center")
            ws[G_colom].alignment = Alignment(horizontal="center")
            ws[H_colom].alignment = Alignment(horizontal="center")
            ws[I_colom].alignment = Alignment(horizontal="center")
            ws[J_colom].alignment = Alignment(horizontal="center")
            ws[K_colom].alignment = Alignment(horizontal="center")

        side = Side(style='medium', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        for row in range(2, r):
            A_rom="A"+str(row)
            B_rom="B"+str(row)
            C_rom="C"+str(row)
            D_rom="D"+str(row)
            E_rom="E"+str(row)
            F_rom="F"+str(row)
            G_rom="G"+str(row)
            H_rom="H"+str(row)
            I_rom="I"+str(row)
            J_rom="J"+str(row)
            K_rom="K"+str(row)
            ws[A_rom].border=border
            ws[B_rom].border=border
            ws[C_rom].border=border
            ws[D_rom].border=border
            ws[E_rom].border=border
            ws[F_rom].border=border
            ws[G_rom].border=border
            ws[H_rom].border=border
            ws[I_rom].border=border
            ws[J_rom].border=border
            ws[K_rom].border=border

      #ファイルを保存する
    wb.save(filename=xlsx_filename)
    #ファイルを閉じる
    wb.close()
