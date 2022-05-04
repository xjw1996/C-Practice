import os
import csv
from importlib.resources import path
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

path = r'D:\\xue_workspace\\DataOfUniversityScore\\ExcelPythonTest\\20220204\\文科二本\\'
path_="D:\\xue_workspace\\DataOfUniversityScore\\ExcelPythonTest\\20220204\\文科二本\\"
filenames = os.listdir(path)
for filename_ in filenames:
    #print(filename)
    str_arr = filename_.split('.')
    #去掉csv后缀的名字
    name_=str_arr[0]
    #print(name_)
    xlsx_filename = "D:\\xue_workspace\\DataOfUniversityScore\\ExcelPythonTest\\20220204\\20212022合并\\"+name_+".xlsx"
    csv_filename=path_+filename_
    font = Font(name='宋体',size=18)
    font2 = Font(bold=True)

    side = Side(style='medium', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)

    wb = openpyxl.Workbook()
    wb.save("D:\\xue_workspace\\DataOfUniversityScore\\ExcelPythonTest\\20220204\\20212022合并\\"+name_+".xlsx")
    telephone="(17790010265,18088685626电话微信)"
    #CSVファイルを開く
    with open(csv_filename,newline="") as csvf:
        #CSVファイルを読み込む
        data=csv.reader(csvf)
        #Excelファイルを開く
        wb=openpyxl.load_workbook(xlsx_filename)
        #sheetを読み込む
        ws=wb["Sheet"]
        #ws.merge_cells(range_string="A1:k1")
        
        r=2
        NumberOf2020 = 0
        NumberOf2020_= 0
        Province2020 = "吉林"
        chinaCollgename = "未定"
        TF_whether=0
        for line in data: 
                c=1
                for v in line:
                    ws.cell(row=r,column=c).value=v
                    c += 1
                    if ws.cell(row=r,column=1).value == "2020":
                        NumberOf2020=r     
                # print(r)
                NumberOf2020_=r
                r += 1

        # 文字列を数値に変換
        TF_whether=int(ws.cell(row=NumberOf2020,column=2).value)
        chinaCollgename=ws.cell(row=3,column=2).value
        # print("TF_whether的值为%d"% TF_whether)

        
        print("大学为%s"% chinaCollgename)
        move_row_number=-(NumberOf2020)+2
        FinalNumebrOf2020=NumberOf2020_-NumberOf2020
        print("Province2020的值为%d"% FinalNumebrOf2020)
        collegename=ws.cell(row=3,column=2).value
        ws['A1'] = collegename+"文科二本2021年在吉专业分数及位次"
        ws['A1'].font = font
        ws = wb.worksheets[0]
        ws["A1"].alignment = Alignment(horizontal="center")
        #合并同类项先对目标区域赋值，在调用此函数
        ws.merge_cells(range_string="A1:k1")
        if TF_whether==1:
            ws['N2'] ='省份'
            ws['O2'] ='年份'
            ws['P2'] ='类型'
            ws['Q2'] ='批次'
            ws['R2'] ='院校'
            ws['S2'] ='专业'
            ws['T2'] ='录取数'
            ws['U2'] ='最低分'
            ws['V2'] ='最低位次'


            ws['N2'].font = font2 
            ws['O2'].font = font2 
            ws['P2'].font = font2 
            ws['Q2'].font = font2
            ws['R2'].font = font2 
            ws['S2'].font = font2
            ws['T2'].font = font2
            ws['U2'].font = font2 
            ws['V2'].font = font2 

            
            ws.move_range("A"+str(NumberOf2020+1)+':'+"I"+str(NumberOf2020_), rows=move_row_number, cols=13, translate=True)
            ws['N1'] = collegename+"文科二本2020年在吉专业分数及位次"
            ws['N1'].font = font
            ws = wb.worksheets[0]
            ws["N1"].alignment = Alignment(horizontal="center")
            ws.merge_cells(range_string="N1:V1")

            for num in range(2, (FinalNumebrOf2020+3)):
                N_colom="N"+str(num)
                O_colom="O"+str(num)
                P_colom="P"+str(num)
                Q_colom="Q"+str(num)
                R_colom="R"+str(num)
                S_colom="S"+str(num)
                T_colom="T"+str(num)
                U_colom="U"+str(num)
                V_colom="V"+str(num)
                ws[N_colom].border=border
                ws[O_colom].border=border
                ws[P_colom].border=border
                ws[Q_colom].border=border
                ws[R_colom].border=border
                ws[S_colom].border=border
                ws[T_colom].border=border
                ws[U_colom].border=border
                ws[V_colom].border=border

                

            for num in range(2, (FinalNumebrOf2020+3)):
                N_colom="N"+str(num)
                O_colom="O"+str(num)
                P_colom="P"+str(num)
                T_colom="T"+str(num)
                U_colom="U"+str(num)
                V_colom="V"+str(num)
                ws[N_colom].alignment = Alignment(horizontal="center")
                ws[O_colom].alignment = Alignment(horizontal="center")
                ws[P_colom].alignment = Alignment(horizontal="center")
                ws[T_colom].alignment = Alignment(horizontal="center")
                ws[U_colom].alignment = Alignment(horizontal="center")
                ws[V_colom].alignment = Alignment(horizontal="center")

            ws['N3'] ="吉林"

        #print(r)
        ws['A2'] ='院校代码'
        ws['F2'].alignment = Alignment(horizontal="center")
        ws['G2'].alignment = Alignment(horizontal="center")
        ws['H2'].alignment = Alignment(horizontal="center")
        ws['I2'].alignment = Alignment(horizontal="center")
        ws['J2'].alignment = Alignment(horizontal="center")
        ws['k2'].alignment = Alignment(horizontal="center")




        ws['A2'].font = font2
        ws['B2'].font = font2
        ws['C2'].font = font2
        ws['D2'].font = font2
        ws['E2'].font = font2
        ws['F2'].font = font2
        ws['G2'].font = font2
        ws['H2'].font = font2
        ws['I2'].font = font2
        ws['J2'].font = font2
        ws['K2'].font = font2


       
        for num in range(3, (NumberOf2020)):
            A_colom="A"+str(num)
            F_colom="F"+str(num)
            G_colom="G"+str(num)
            H_colom="H"+str(num)
            I_colom="I"+str(num)
            J_colom="J"+str(num)
            K_colom="K"+str(num)
            ws[A_colom].alignment = Alignment(horizontal="center")
            ws[F_colom].alignment = Alignment(horizontal="center")
            ws[G_colom].alignment = Alignment(horizontal="center")
            ws[H_colom].alignment = Alignment(horizontal="center")
            ws[I_colom].alignment = Alignment(horizontal="center")
            ws[J_colom].alignment = Alignment(horizontal="center")
            ws[K_colom].alignment = Alignment(horizontal="center")

        
        for row in range(2, (NumberOf2020)):
            A_rom="A"+str(row)
            B_rom="B"+str(row)
            C_rom="C"+str(row)
            D_rom="D"+str(row)
            E_rom="E"+str(row)
            F_rom="F"+str(row)
            G_rom="G"+str(row)
            H_rom="H"+str(row)
            I_rom="I"+str(row)
            J_rom="J"+str(row)
            K_rom="K"+str(row)
            ws[A_rom].border=border
            ws[B_rom].border=border
            ws[C_rom].border=border
            ws[D_rom].border=border
            ws[E_rom].border=border
            ws[F_rom].border=border
            ws[G_rom].border=border
            ws[H_rom].border=border
            ws[I_rom].border=border
            ws[J_rom].border=border
            ws[K_rom].border=border

        



        tel_colom3="A"+str(NumberOf2020)
        tel_colom4="B"+str(NumberOf2020)
        tel_colom="C"+str(NumberOf2020)
        tel_colom2="I"+str(NumberOf2020)
        tel_font = Font(name='宋体',size=15,color='FF0000')
        ws[tel_colom]=telephone
        ws[tel_colom3]=" "
        ws[tel_colom4]=" "
        ws[tel_colom].font = tel_font
        ws[tel_colom].alignment = Alignment(horizontal="center")
        ws.merge_cells(range_string="C"+str(NumberOf2020)+":"+"I"+str(NumberOf2020))

    #ファイルを保存する
    wb.save(filename=xlsx_filename)
    #ファイルを閉じる
    wb.close()




 #https://gammasoft.jp/blog/text-center-alignment-with-openpyxl/
 #https://pg-chain.com/python-excel-cell-write
 #https://tonari-it.com/python-openpyxl-font/
 #https://blog.csdn.net/qq_43342294/article/details/109356948
 #python 文字列の中に変数を挿入する　https://qiita.com/niwasawa/items/27641b803db31f93b8e6
 #openpyxl  で枠線を引く　https://qiita.com/github-nakasho/items/358e5602aeda81c58c81
