import os
import csv
from importlib.resources import path
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

# path = 'C:\\Users\\jinue\\Desktop\\数据\\理科二本.csv'
# path_="D:\\xue_workspace\\DataOfUniversityScore\\文科二批大学\\"
# filenames = os.listdir(path)

xlsx_filename = "C:\\Users\\jinue\\Desktop\\数据\\20200428\\文科二本\\2021-2019文科二本院校录取位次及位次差"+".xlsx"
csv_filename="C:\\Users\\jinue\\Desktop\\数据\\20200428\\文科二本\\2021-2019理科一本院校录取位次及位次差.csv"
font = Font(name='宋体',size=18)

wb = openpyxl.Workbook()
wb.save("C:\\Users\\jinue\\Desktop\\数据\\20200428\\文科二本\\2021-2019文科二本院校录取位次及位次差"+".xlsx")
#CSVファイルを開く
with open(csv_filename,newline="") as csvf:
    #CSVファイルを読み込む
    data=csv.reader(csvf)
    #Excelファイルを開く
    wb=openpyxl.load_workbook(xlsx_filename)
    #sheetを読み込む
    ws=wb["Sheet"]
    ws['A1'] = "2021-2019文科二本院校录取位次及位次差"
    ws['A1'].font = font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(range_string="A1:Q1")
    
    r=2
    for line in data: 
        c=1
        for v in line:
            ws.cell(row=r,column=c).value=v
            c += 1
        r += 1
    #print(r)
    ws['A3'] ='0'
    ws['A2'] ='序号'
    ws['B2'] ='院校代码'
    ws['C2'] ='院校名称'
    ws['D2'] ='备注'
    ws['E2'] ='批次'
    ws['F2'] ='2021投档线'
    ws['G2'] ='2021计划数'
    ws['H2'] ='2021位次'
    ws['I2'] ='2020位次'
    ws['J2'] ='位次差'
    ws['k2'] ='2019位次'
    ws['L2'] ='985'
    ws['M2'] ='211'
    ws['N2'] ='双一流'
    ws['O2'] ='城市'
    ws['P2'] ='省份'
    ws['Q2'] ='性质'
    
    for num in range(3, r):
        A_colom="A"+str(num)
        B_colom="B"+str(num)
        E_colom="E"+str(num)
        F_colom="F"+str(num)
        G_colom="G"+str(num)
        H_colom="H"+str(num)
        I_colom="I"+str(num)
        J_colom="J"+str(num)
        K_colom="K"+str(num)
        L_colom="L"+str(num)
        M_colom="M"+str(num)
        N_colom="N"+str(num)
        O_colom="O"+str(num)
        P_colom="P"+str(num)
        Q_colom="Q"+str(num)
        ws[A_colom].alignment = Alignment(horizontal="center")
        ws[B_colom].alignment = Alignment(horizontal="center")
        ws[E_colom].alignment = Alignment(horizontal="center")
        ws[F_colom].alignment = Alignment(horizontal="center")
        ws[G_colom].alignment = Alignment(horizontal="center")
        ws[H_colom].alignment = Alignment(horizontal="center")
        ws[I_colom].alignment = Alignment(horizontal="center")
        ws[J_colom].alignment = Alignment(horizontal="center")
        ws[K_colom].alignment = Alignment(horizontal="center")
        ws[L_colom].alignment = Alignment(horizontal="center")
        ws[M_colom].alignment = Alignment(horizontal="center")
        ws[N_colom].alignment = Alignment(horizontal="center")
        ws[O_colom].alignment = Alignment(horizontal="center")
        ws[P_colom].alignment = Alignment(horizontal="center")
        ws[Q_colom].alignment = Alignment(horizontal="center")

    # side = Side(style='medium', color='000000')
    # border = Border(top=side, bottom=side, left=side, right=side)
    # for row in range(2, r):
    #     A_rom="A"+str(row)
    #     B_rom="B"+str(row)
    #     C_rom="C"+str(row)
    #     D_rom="D"+str(row)
    #     E_rom="E"+str(row)
    #     F_rom="F"+str(row)
    #     G_rom="G"+str(row)
    #     H_rom="H"+str(row)
    #     I_rom="I"+str(row)
    #     J_rom="J"+str(row)
    #     K_rom="K"+str(row)
    #     ws[A_rom].border=border
    #     ws[B_rom].border=border
    #     ws[C_rom].border=border
    #     ws[D_rom].border=border
    #     ws[E_rom].border=border
    #     ws[F_rom].border=border
    #     ws[G_rom].border=border
    #     ws[H_rom].border=border
    #     ws[I_rom].border=border
    #     ws[J_rom].border=border
    #     ws[K_rom].border=border
#ファイルを保存する
wb.save(filename=xlsx_filename)
#ファイルを閉じる
wb.close()




 #https://gammasoft.jp/blog/text-center-alignment-with-openpyxl/
 #https://pg-chain.com/python-excel-cell-write
 #https://tonari-it.com/python-openpyxl-font/
 #https://blog.csdn.net/qq_43342294/article/details/109356948
 #python 文字列の中に変数を挿入する　https://qiita.com/niwasawa/items/27641b803db31f93b8e6
 #openpyxl  で枠線を引く　https://qiita.com/github-nakasho/items/358e5602aeda81c58c81
