from cgi import test
import os
import csv
from importlib.resources import path
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
sql_dir_war="C:\\Users\\jinue\\Desktop\\薛经纬1\\薛经纬1\\理科一批不同段位报考指南2"
dbtype_list = os.listdir(sql_dir_war)
csv_filename="C:\\Users\\jinue\\Desktop\\测试.xlsx"

arr=[]
row_num=1
for dbtype in dbtype_list:
    # if os.path.isfile(os.path.join(sql_dir_war,dbtype)):
    #     #dbtype_list.remove(dbtype)
    arr.append(dbtype)

# for n in range(len(arr)):
#     print(arr[n])
# print(len(dbtype_list))
wb=openpyxl.load_workbook(csv_filename)
ws=wb["Sheet1"]
for num in range(len(dbtype_list)):
    A_colom="A"+str(num+1)
    colom_num=2
    with open(csv_filename,newline="") as csvf:
        ws[A_colom]=arr[num]
        path="C:\\Users\\jinue\\Desktop\\薛经纬1\\薛经纬1\\理科一批不同段位报考指南2\\"+arr[num]
        print(path)
        filenames = os.listdir(path)
        for filename_ in filenames:
            ws.cell(row=row_num,column=colom_num).value =filename_
            colom_num=colom_num+1
            print(filename_)
        print("--------------------------------------")
        row_num+=1
wb.save(filename="C:\\Users\\jinue\\Desktop\\测试.xlsx")
wb.close()