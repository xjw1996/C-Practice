import os
import csv
from importlib.resources import path
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

path = r'D:\\xue_workspace\\DataOfUniversityScore\\文科二批大学'
path_="D:\\xue_workspace\\DataOfUniversityScore\\文科二批大学\\"
filenames = os.listdir(path)
for filename_ in filenames:
    #print(filename)
    str_arr = filename_.split('.')
    #去掉csv后缀的名字
    name_=str_arr[0]
    #print(name_)
    xlsx_filename = "D:\\xue_workspace\\DataOfUniversityScore\\ExcelPythonTest\\Collge\\文科二本大学\\"+name_+".xlsx"
    csv_filename=path_+filename_
    font = Font(name='宋体',size=18)
    
    wb = openpyxl.Workbook()
    wb.save("D:\\xue_workspace\\DataOfUniversityScore\\ExcelPythonTest\\Collge\\文科二本大学\\"+name_+".xlsx")
    telephone="(17790010265,18088685626电话微信)"
    #CSVファイルを開く
    with open(csv_filename,newline="") as csvf:
        #CSVファイルを読み込む
        data=csv.reader(csvf)
        #Excelファイルを開く
        wb=openpyxl.load_workbook(xlsx_filename)
        #sheetを読み込む
        ws=wb["Sheet"]
        #ws.merge_cells(range_string="A1:k1")
        
        ws['A1'] = name_
        ws['A1'].font = font
        ws = wb.worksheets[0]
        ws["A1"].alignment = Alignment(horizontal="center")
        #合并同类项先对目标区域赋值，在调用此函数
        ws.merge_cells(range_string="A1:k1")
        r=2
        for line in data: 
            c=1
            for v in line:
                ws.cell(row=r,column=c).value=v
                c += 1
            r += 1
        #print(r)
        ws['A2'] ='院校代码'
        ws['F2'].alignment = Alignment(horizontal="center")
        ws['G2'].alignment = Alignment(horizontal="center")
        ws['H2'].alignment = Alignment(horizontal="center")
        ws['I2'].alignment = Alignment(horizontal="center")
        ws['J2'].alignment = Alignment(horizontal="center")
        ws['k2'].alignment = Alignment(horizontal="center")
        for num in range(3, r):
            A_colom="A"+str(num)
            F_colom="F"+str(num)
            G_colom="G"+str(num)
            H_colom="H"+str(num)
            I_colom="I"+str(num)
            J_colom="J"+str(num)
            K_colom="K"+str(num)
            ws[A_colom].alignment = Alignment(horizontal="center")
            ws[F_colom].alignment = Alignment(horizontal="center")
            ws[G_colom].alignment = Alignment(horizontal="center")
            ws[H_colom].alignment = Alignment(horizontal="center")
            ws[I_colom].alignment = Alignment(horizontal="center")
            ws[J_colom].alignment = Alignment(horizontal="center")
            ws[K_colom].alignment = Alignment(horizontal="center")

        side = Side(style='medium', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        for row in range(2, r):
            A_rom="A"+str(row)
            B_rom="B"+str(row)
            C_rom="C"+str(row)
            D_rom="D"+str(row)
            E_rom="E"+str(row)
            F_rom="F"+str(row)
            G_rom="G"+str(row)
            H_rom="H"+str(row)
            I_rom="I"+str(row)
            J_rom="J"+str(row)
            K_rom="K"+str(row)
            ws[A_rom].border=border
            ws[B_rom].border=border
            ws[C_rom].border=border
            ws[D_rom].border=border
            ws[E_rom].border=border
            ws[F_rom].border=border
            ws[G_rom].border=border
            ws[H_rom].border=border
            ws[I_rom].border=border
            ws[J_rom].border=border
            ws[K_rom].border=border
        tel_colom_num=r-1
        tel_colom="C"+str(r)
        tel_colom2="I"+str(r)
        tel_font = Font(name='宋体',size=14,color='FF0000')
        ws[tel_colom]=telephone
        ws[tel_colom].font = tel_font
        ws[tel_colom].alignment = Alignment(horizontal="center")
        ws.merge_cells(range_string="C"+str(r)+":"+"I"+str(r))

    #ファイルを保存する
    wb.save(filename=xlsx_filename)
    #ファイルを閉じる
    wb.close()




 #https://gammasoft.jp/blog/text-center-alignment-with-openpyxl/
 #https://pg-chain.com/python-excel-cell-write
 #https://tonari-it.com/python-openpyxl-font/
 #https://blog.csdn.net/qq_43342294/article/details/109356948
 #python 文字列の中に変数を挿入する　https://qiita.com/niwasawa/items/27641b803db31f93b8e6
 #openpyxl  で枠線を引く　https://qiita.com/github-nakasho/items/358e5602aeda81c58c81
